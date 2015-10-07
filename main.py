# -*- coding: cp1251 -*-
from openpyxl import *
import xlwt

#��������� ���� �� ������ � �������
def HochladenInfo(files):
    data = {}
    for name in files:
        data[name]=[]
        wb = load_workbook(filename = name, use_iterators = True)
        ws = wb.worksheets[0]
        rowcount = 0
        for row in ws.iter_rows():
            data[name].append([])    
            for cell in row:
                data[name][rowcount].append(cell.value)
            rowcount +=1
        print 'File "'+name+'" ist hochgeladen'
        
    return data

#������ ��������
def Parsing_usw(data):
    for a_rownum in range(len(data['AddOrEddit.xlsx'])):
        a_cell = data['AddOrEddit.xlsx'][a_rownum][0]
        MnemonikaIsFound = False
        
        for m_rownum in range(len(data['mnemonics.xlsx'])):
            m_cell = data['mnemonics.xlsx'][m_rownum][1]
            if (a_cell==m_cell): # ��������� �� �������� ��������� � �� ������� � ��
                MnemonikaIsFound = True
                Ubersetzung(data['mnemonics.xlsx'][m_rownum][0], a_rownum, data)
                break

        if ((MnemonikaIsFound == False) & (a_cell!=None)):
            AddMnemonik(a_cell, data)
            Ubersetzung(data['mnemonics.xlsx'][-1][0], a_rownum, data)
            break

#���������� ����� ���������
def AddMnemonik(name, data):
    last_pos = data['mnemonics.xlsx'][-1][0] + 1
    data['mnemonics.xlsx'].append([])
    data['mnemonics.xlsx'][-1].append(last_pos)#��������� ID � ���������
    data['mnemonics.xlsx'][-1].append(name)#���
    data['mnemonics.xlsx'][-1].append(1)#Area
    #print "----------\n��������� ����� ���������:", name," | � id -> ", data['mnemonics.xlsx'][-1][0],"\n----------"
            
#���������� ����� ���������        
def Hinzufugen(languages, id_m, a_rownum, data):
    while len(languages)>0:
        last_pos = data['translation.xlsx'][-1][0] + 1
        data['translation.xlsx'].append([])
        data['translation.xlsx'][-1].append(last_pos)#��������� ID � ��������
        data['translation.xlsx'][-1].append(id_m)#��������� ID ���������
        data['translation.xlsx'][-1].append(languages[0])#��� �����
        data['translation.xlsx'][-1].append(data['AddOrEddit.xlsx'][a_rownum][languages[0]+1])#�������

        #print "�������� ����� ������� ��� ��������� � id ->",id_m,"| �� �����", languages[0], "| Ubersetzung: ",data['translation.xlsx'][-1][3]
        del languages[0]
        
#����� ������������ � �� ����������� (���� �����)
def Ubersetzung(id_m, a_rownum, data):
    languages = [0,1,2] #0 - ���, 1- ���, 2-����
    while languages: # ���� �� ����� 3 ���������
        for t_rownum in range(len(data['translation.xlsx'])):
            x=data['translation.xlsx'][t_rownum][2]
            if x > 2: # ������ ��� �� ������� (� �������. ��� ��������)
                continue
            #���� �����������, ������� �� ��� (� �� ��������), ���� �� ����� ������� � AddOrEdit, ���� ���� �� ��������.
            if ((id_m==data['translation.xlsx'][t_rownum][1])  and (data['AddOrEddit.xlsx'][a_rownum][x+1] != None) ):
                data['translation.xlsx'][t_rownum][3] = data['AddOrEddit.xlsx'][a_rownum][x+1] # ���� ���� ������ ���, ��������� �� ������ � AddOrEddit........����� ����� ���� ������� ������ :(
                #print "�������� ������� �� ��������� � id -> ",id_m,"| �� �����= ",x
                languages.remove(x) # ������� ������/��������, �������.
                continue
            
        Hinzufugen(languages, id_m, a_rownum, data)
        
#��������� � ��������� excel
def Speichern(data, name=[]):
    counter = 0
    for doc in name:
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("MyData")
        for row_index, row in enumerate(data[name[counter]]):
            for col_index, cell_value in enumerate(row):
                sheet.write(row_index, col_index, cell_value)
                #print(row_index, col_index, cell_value)
        workbook.save("neu_"+name[counter][:-1])
        print "neu_"+name[counter], " <- ist saved"
        counter+=1
        
        
def main():
    dic = {}
    dic = HochladenInfo(['mnemonics.xlsx', 'translation.xlsx', 'AddOrEddit.xlsx'])
    print "\nThe documents were downloaded!"
    Parsing_usw(dic)
    print "\nThe documents is ready!\n"
    Speichern(dic, ['mnemonics.xlsx','translation.xlsx'])
    print "\nThe Job is dont!\n"
    raw_input("Press the enter key to exit.")
    print "Have a nice day!"
    exit()
    
if __name__ == "__main__":
    main()
